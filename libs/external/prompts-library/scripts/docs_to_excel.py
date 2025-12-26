#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docs_to_excel.py

Documents → Excel converter: rebuild a workbook from prompts folders.

Rules (per STRUCTURE_AND_CONVERSION_SPEC.md):
- Each folder under prompt-library/prompts that matches "(N)_<name>" or any folder is a sheet
- For each file matching "(r,c)_*.md", write its full text to Excel cell (r,c), 1-based
- Title part in filename is ignored for cell value
- Non-matching files are ignored
- Optionally clears existing workbook or merges (default: overwrite generate new)

Usage:
  python prompt-library/scripts/docs_to_excel.py --out "rebuilt.xlsx"
  # optional: --prompts-dir prompt-library/prompts --clear
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, Tuple

import pandas as pd
from openpyxl import Workbook

FOLDER_PREFIX_RE = re.compile(r"^\((\d+)\)_")
FILE_NAME_RE = re.compile(r"^\((\d+),(\d+)\)_.*\.md$")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Rebuild Excel workbook from prompt folders")
    p.add_argument("--prompts-dir", type=str, default="prompt-library/prompts", help="Prompts root directory")
    p.add_argument("--out", type=str, required=True, help="Output Excel file path")
    return p.parse_args()


def list_sheet_folders(prompts_root: Path) -> Dict[str, Path]:
    sheets: Dict[str, Path] = {}
    for child in sorted(prompts_root.iterdir()):
        if not child.is_dir():
            continue
        if child.name == "prompt-category":
            # legacy; skip auto-generated category
            continue
        sheets[child.name] = child
    return sheets


def extract_rc(name: str) -> Tuple[int, int] | None:
    m = FILE_NAME_RE.match(name)
    if not m:
        return None
    r = int(m.group(1))
    c = int(m.group(2))
    return r, c


def main() -> None:
    args = parse_args()
    prompts_root = Path(args.prompts_dir).resolve()
    out_path = Path(args.out).resolve()

    if not prompts_root.exists():
        raise FileNotFoundError(f"Prompts directory not found: {prompts_root}")

    sheet_folders = list_sheet_folders(prompts_root)
    if not sheet_folders:
        raise RuntimeError("No sheet folders found under prompts root")

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    for folder_name, folder_path in sheet_folders.items():
        # Recover original sheet name (try to drop ordering prefix "(N)_")
        m = FOLDER_PREFIX_RE.match(folder_name)
        sheet_name = folder_name[m.end():] if m else folder_name
        if not sheet_name:
            sheet_name = folder_name
        ws = wb.create_sheet(title=sheet_name)

        # Aggregate cells
        max_row = 0
        max_col = 0
        cells: Dict[Tuple[int, int], str] = {}
        for file in folder_path.iterdir():
            if not file.is_file() or not file.name.endswith('.md'):
                continue
            rc = extract_rc(file.name)
            if not rc:
                continue
            r, c = rc
            text = file.read_text(encoding='utf-8')
            # Trim a single trailing newline for cell value aesthetics
            if text.endswith("\n"):
                text = text[:-1]
            cells[(r, c)] = text
            if r > max_row:
                max_row = r
            if c > max_col:
                max_col = c

        # Write into sheet
        for (r, c), val in cells.items():
            ws.cell(row=r, column=c, value=val)

    # Save workbook
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"✅ Rebuilt Excel saved to: {out_path}")


if __name__ == "__main__":
    main()
